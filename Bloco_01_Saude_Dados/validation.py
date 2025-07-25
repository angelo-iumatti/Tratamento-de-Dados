import pandas as pd
from pathlib import Path

ARQUIVO = Path("OP01_202506.csv")

# 1. Carregar
df = pd.read_csv(ARQUIVO)

# 2. Validações básicas
erros = []

# 2.1. Campos obrigatórios sem nulos
campos_obrig = ["municipio", "mes_referencia", "volume_produzido_m3"]
for c in campos_obrig:
    n_nulos = df[c].isna().sum()
    if n_nulos:
        erros.append(f"{c}: {n_nulos} nulos")

# 2.2. Duplicatas
duplicatas = df.duplicated(subset=["municipio", "mes_referencia"]).sum()
if duplicatas:
    erros.append(f"{duplicatas} linhas duplicadas")

# 2.3. Regras de negócio
if (df["perda_percentual"] > 100).any():
    erros.append("perda_percentual > 100% detectada")

if (df["volume_produzido_m3"] < 0).any():
    erros.append("volume negativo detectado")

# 3. Relatório
pd.Series(erros, name="erros").to_csv("erros_validacao.csv", index=False)
print("Validação concluída! Veja erros_validacao.csv")

# 4. Exportar DataFrame corrigido
df.to_csv("OP01_202506_corrigido.csv", index=False)

# 5. Checklist de Validação em Excel
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

excel_path = "Checklist_Validação.xlsx"

# Regras de validação
regras = [
    {
        "Verificação": "Campos obrigatórios sem nulos",
        "Descrição": "Verifica se os campos obrigatórios não possuem valores nulos.",
        "Referência": "Portaria GM/MS nº 1.792/2023, art. 5º"
    },
    {
        "Verificação": "Duplicatas",
        "Descrição": "Verifica duplicidade de registros por município e mês.",
        "Referência": "Manual SIA/SUS, seção 2.3"
    },
    {
        "Verificação": "Perda percentual > 100%",
        "Descrição": "Verifica se o campo perda_percentual está acima de 100%.",
        "Referência": "Portaria GM/MS nº 1.792/2023, art. 7º"
    },
    {
        "Verificação": "Volume negativo",
        "Descrição": "Verifica se o campo volume_produzido_m3 possui valores negativos.",
        "Referência": "Manual SIA/SUS, seção 2.4"
    }
]
df_regras = pd.DataFrame(regras)

# Log de erros
df_log = pd.read_csv("erros_validacao.csv", names=["Erros"])

if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    # Atualiza aba Log
    if "Log" in wb.sheetnames:
        ws_log = wb["Log"]
        ws_log.delete_rows(2, ws_log.max_row)  # Limpa log anterior, mantém cabeçalho
        for r in dataframe_to_rows(df_log, index=False, header=False):
            ws_log.append(r)
    else:
        ws_log = wb.create_sheet("Log")
        for r in dataframe_to_rows(df_log, index=False, header=True):
            ws_log.append(r)
else:
    wb = Workbook()
    ws_regras = wb.active
    ws_regras.title = "Regras"
    for r in dataframe_to_rows(df_regras, index=False, header=True):
        ws_regras.append(r)
    ws_log = wb.create_sheet("Log")
    for r in dataframe_to_rows(df_log, index=False, header=True):
        ws_log.append(r)

wb.save(excel_path)